import pandas as pd
from openpyxl import load_workbook

from func import httml_soup, source_link, next_page
from openpyxl.styles import PatternFill


pages_domen = 'https://autopiter.ru/'

workbook = load_workbook('список.xlsx')

file_path = 'список.xlsx'

sheet = workbook.active

df = pd.read_excel(file_path)



def pars_links_check():
    print('Введите какое количество ссылок нужно проверить?')
    print('Нажмите просто Enter чтобы выполнить всё')

    try:
        counter = int(input())
    except ValueError:
        counter = None

    sheet['G1'] = 'ссылка'

    workbook.save('список.xlsx')

    s = 0
    # Перебор всех строк в exel
    for index, row in df.iterrows():
        s += 1
        d =row.to_dict()
        artlec = str(d['Артикул']).replace(' ', '%20').replace('/', '%2F').strip()
        page = f'{pages_domen}goods/{artlec}'
        soup = httml_soup(page)
        #получаем ссылку на интересующий нас подшипник
        ass = source_link(soup, d['Бренд'].strip(), str(d['Артикул']).strip())
        if ass is None:
            list_links = next_page(soup)
            if list_links:
                for i in list_links:
                    print('Переходим к след странице по номеру')
                    soup1 = httml_soup(pages_domen + i)
                    ass = source_link(soup1, d['Бренд'], d['Артикул'])
                    if ass:
                        break

        if ass == d['ссылка']:
            print(f"Ссылка {index + 2}, {ass}, {d['Бренд']}, {d['Артикул']} совпадает")
        else:
            if pd.isna(d['ссылка']):
                sheet['G' + str(index + 2)] = ass
                print(f"{index + 2} ссылка не обнаружена, ячейка пустая устонавливаем ссылку {ass}")
            else:
                str_log = f"Ссылка {index + 2},  {ass}, {d['Бренд']}, артикул {d['Артикул']} НЕ СОВПАЛА СО СТАРОЙ {d['ссылка']}"
                sheet['G' + str(index + 2)] = ass
                red_fill = PatternFill(
                            start_color='FF0000',
                            end_color='FF0000',
                            fill_type='solid'
                            )
                        # Применяем заливку к ячейке
                sheet['G' + str(index + 2)].fill = red_fill
                print(str_log)
        if index % 10 == 0:
            workbook.save('список.xlsx')
        if counter:
            if s == counter:
                break

    workbook.save('список.xlsx')