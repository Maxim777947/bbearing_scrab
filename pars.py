from bs4 import BeautifulSoup
from selenium import webdriver
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from func import httml_soup, source_link, next_page, price, price_min


start = time.perf_counter()

workbook = load_workbook('список.xlsx')

sheet = workbook.active

file_path = 'список.xlsx'

df = pd.read_excel(file_path)

pages_domen = 'https://autopiter.ru/'

pages_domen = 'https://autopiter.ru/'

mehanica = [30398, 32893]

sheet['G1'] = 'ссылка'
sheet['H1'] = 'Уточнено'

workbook.save('список.xlsx')

print("Ведите:")
print('1 - Запуск поиска ссылок на подшипники и запись их в таблицу Exel ')
print('2 - Запуск поиска минимальной цены и запись в таблицу Exel' )
f = int(input())




if f == 1:
    print('Введите какое количество ссылок нужно создать?')
    print('Нажмите просто Enter чтобы выполнить всё')
    try:
        counter = int(input())
    except ValueError:
        counter = None

    s = 0
    # Перебор всех строк в exel
    for index, row in df.iterrows():
        d =row.to_dict()
        if pd.isna(d['ссылка']):
            artlec = str(d['Артикул']).replace(' ', '%20').replace('/', '%2F').strip()
            page = f'{pages_domen}goods/{artlec}'
            soup = httml_soup(page)
            #получаем ссылку на интересующий нас подшипник
            ass = source_link(soup, d['Бренд'].strip(), str(d['Артикул']).strip())
            if ass is None:
                list_links = next_page(soup)
                if list_links:
                    for i in list_links:
                        print('Переходим к след странице по номеру')
                        soup1 = httml_soup(pages_domen + i)
                        ass = source_link(soup1, d['Бренд'], d['Артикул'])
                        if ass:
                            break
            if ass:
                sheet['G' + str(index + 2)] = ass
                str_log = f"{index + 2} добавление ссылки {ass} на подшипник бренда {d['Бренд']}, артикул {d['Артикул']}"
                print(str_log)
                s += 1
        if index%5 == 0:
            workbook.save('список.xlsx')
        if counter:
            if s == counter:
                break
    workbook.save('список.xlsx')

elif f == 2:
    print('Введите какое количество подшипников нужно проверить?')
    print('Нажмите просто Enter чтобы выполнить всё')
    try:
        counter = int(input())
    except ValueError:
        counter = None
    s = 0
    browser = webdriver.Chrome()

    for i in range(5):
        print('НЕ ЗАБУДЬ РАЗВЕРНУТЬ КАРТОЧКУ ТОВАРА')

    for index, row in df.iterrows():
        d =row.to_dict()
        if pd.isna(d['Уточнено']):
            browser.get('https://autopiter.ru' + str(d['ссылка']))
            soup = BeautifulSoup(browser.page_source, "lxml")
            new_price = price(soup)
            price_m = price_min(soup, new_price)
            if new_price is None:
                print('new_price = None нажми я не робот')
                time.sleep(3)
            if new_price is not None:
                sheet['H' + str(index + 2)] = new_price
                s += 1
                print(index + 2, f'обновление цены на {new_price}')
            if price_m and new_price:
                if price_m['priceId'] not in mehanica:
                    print('магазин не наш')
                    red_fill = PatternFill(
                        start_color='FF0000',
                        end_color='FF0000',
                        fill_type='solid'
                        )
                    # Применяем заливку к ячейке
                    sheet['H' + str(index + 2)].fill = red_fill
        if index%5 == 0:
            workbook.save('список.xlsx')
        new_price = None

        if counter:
            if s == counter:
                break
    workbook.save('список.xlsx')

    browser.quit()



finish = time.perf_counter()
print('Время работы: ' + str(finish - start))




# очисти ячеуйк
# разверни карточку товара
# файл должен называться список
# каждая ячейка в столбце в которую будет записывается цена должна быть
#  внутере пустой нажми удалить его чтобы компьютер видел что там ничего нету если в ячеки
#  будет хоть один пробел или она будет закрашена и т.д и тп комп будет думать что там что то есть и не будет в эту ячейку ни чего записыват'' 