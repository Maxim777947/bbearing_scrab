import asyncio
import aiohttp
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from func import httml_soup, source_link, next_page, replace_article


PAGES_DOMEN = 'https://autopiter.ru/'

FILE_PATH = 'список.xlsx'


async def fetch(session, d_dict, semaphore):
    '''Функция ищет ссылку на подшипник'''
    async with semaphore:
        async with session.get(d_dict['url']) as response:
            html_content = await response.text()
            soup = BeautifulSoup(html_content, 'html.parser')
            ass = source_link(soup, d_dict['brend'], d_dict['article'])
            ss_tuple = (d_dict['index'] + 2, ass)
            print(f"Найдена ссылка {ss_tuple}")
            return ss_tuple


async def parse_links(counter=None, filter=False):
    '''1313'''
    workbook = load_workbook('список.xlsx')

    sheet = workbook.active

    df = pd.read_excel(FILE_PATH)

    semaphore = asyncio.Semaphore(5)

    async with aiohttp.ClientSession() as session:
        tasks = []
        s = 0
        for index, row in df.iterrows():
            d = row.to_dict()
            brend = str(d['Бренд']).title()
            if pd.isna(d['ссылка']):
                artlec = replace_article(d['Артикул'])
                if filter:
                    url = f'{PAGES_DOMEN}goods/{artlec}?filteredBrands%5B0%5D={brend}'
                else:
                    url = f'{PAGES_DOMEN}goods/{artlec}'
                d_dict = {
                    'index': index,
                    'url': url,
                    'brend': brend,
                    'article': d['Артикул'],
                }
                tasks.append(fetch(session, d_dict, semaphore))
                s += 1
                if counter:
                    if s == counter:
                        break
        # Выполняем все задачи асинхронно
        responses = await asyncio.gather(*tasks)
        # Обработка ответов
        for i in responses:
            if i[1]:
                sheet['G' + str(i[0])] = i[1]
                str_log = f"{i[0]} добавление ссылки {i[1]} в таблицу"
                print(str_log)
        workbook.save('список.xlsx')


def pars_linksssss(counnter):
    s = 0
    '''Функция доробатывает за parse_links'''
    workbook = load_workbook('список.xlsx')

    sheet = workbook.active

    df = pd.read_excel(FILE_PATH)
    for index, row in df.iterrows():
        d =row.to_dict()
        
        if pd.isna(d['ссылка']):
            artlec = replace_article(d['Артикул'])
            page = f'{PAGES_DOMEN}goods/{artlec}'
            soup = httml_soup(page)
            #получаем ссылку на интересующий нас подшипник
            ass = source_link(soup, str(d['Бренд']).strip(), str(d['Артикул']).strip())
            list_links = next_page(soup)
            if ass is None and list_links:
                for i in list_links:
                    print('Переходим к след странице по номеру')
                    llink = PAGES_DOMEN + i.strip('/')
                    soup1 = httml_soup(llink)
                    ass = source_link(soup1, str(d['Бренд'].strip()), str(d['Артикул']).strip())
                    if ass:
                        break
            if ass:
                s += 1
                sheet['G' + str(index + 2)] = ass
                str_log = (f"{index + 2} добавление ссылки "
                          f"{ass} на подшипник бренда "
                          f"{d['Бренд']}, артикул {d['Артикул']}")
                print(str_log)
                
        if s == counnter:
            break
    workbook.save('список.xlsx')
