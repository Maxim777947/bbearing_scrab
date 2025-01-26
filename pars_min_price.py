import time
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from selenium import webdriver
from func import source_price, price_min




workbook = load_workbook('список.xlsx')

sheet = workbook.active

FILE_PATH = 'список.xlsx'

df = pd.read_excel(FILE_PATH)



mehanica = [30398, 32893]

def pars_price():
    '''Основная функция парсит цены и вносит их в таблицу'''
    print('Введите какое количество подшипников нужно проверить?')
    print('Нажмите просто Enter чтобы выполнить всё')
    try:
        counter = int(input())
    except ValueError:
        counter = None
    s = 0

    sheet['H1'] = 'Уточнено'

    browser = webdriver.Chrome()

    for _ in range(5):
        print('НЕ ЗАБУДЬ РАЗВЕРНУТЬ КАРТОЧКУ ТОВАРА')

    for index, row in df.iterrows():
        d =row.to_dict()
        if pd.isna(d['Уточнено']) and pd.notna(d['ссылка']):
            browser.get('https://autopiter.ru' + str(d['ссылка']))
            soup = BeautifulSoup(browser.page_source, "lxml")
            min_price = source_price(soup)
            price_m = price_min(soup, min_price)
            if min_price is None:
                print('min_price = None нажми я не робот')
                time.sleep(2)
            if min_price is not None:
                sheet['H' + str(index + 2)] = min_price
                s += 1
                print(index + 2, f'обновление цены на {min_price}')
            if min_price == 'Предложений по запрошенному номеру не найдено':
                yellow_fill = PatternFill(
                        start_color='FFD700',
                        end_color='FFD700',
                        fill_type='solid'
                        )
                # Применяем заливку к ячейке
                sheet['H' + str(index + 2)].fill = yellow_fill
            if price_m and min_price:
                if price_m['priceId'] not in mehanica:
                    print('магазин не наш')
                    red_fill = PatternFill(
                        start_color='FF0000',
                        end_color='FF0000',
                        fill_type='solid'
                        )
                    # Применяем заливку к ячейке
                    sheet['H' + str(index + 2)].fill = red_fill
        min_price = None

        if counter:
            if s == counter:
                break
    workbook.save('список.xlsx')

    browser.quit()
